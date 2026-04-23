"""Agent 5: REPORTER — fpdf2 PDF 견적서 생성"""
import base64
import logging
import os
from datetime import datetime

logger = logging.getLogger(__name__)

try:
    from fpdf import FPDF
    _FPDF_OK = True
except ImportError:
    _FPDF_OK = False

# 한글 폰트 경로 후보 (regular / bold 분리)
# 번들 폰트를 최우선으로 검색 (Render/Fly.io 등 apt 미지원 환경 대응)
_BUNDLE_FONT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "fonts")
_FONT_REGULAR_CANDIDATES = [
    os.path.join(_BUNDLE_FONT_DIR, "NanumGothic.ttf"),
    os.path.join(_BUNDLE_FONT_DIR, "NanumGothicBold.ttf"),
    "C:/Windows/Fonts/malgun.ttf",
    "C:/Windows/Fonts/NanumGothic.ttf",
    "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
]

_FONT_BOLD_CANDIDATES = [
    os.path.join(_BUNDLE_FONT_DIR, "NanumGothicBold.ttf"),
    os.path.join(_BUNDLE_FONT_DIR, "NanumGothic.ttf"),
    "C:/Windows/Fonts/malgunbd.ttf",
    "C:/Windows/Fonts/NanumGothicBold.ttf",
    "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
]


def _find_font() -> str | None:
    for path in _FONT_REGULAR_CANDIDATES:
        if os.path.exists(path):
            return path
    return None


def _find_bold_font() -> str | None:
    """bold 전용 폰트. regular와 동일 파일이면 None 반환 (충돌 방지)"""
    reg = _find_font()
    for path in _FONT_BOLD_CANDIDATES:
        if os.path.exists(path) and path != reg:
            return path
    return None


def _check_page(pdf: "FPDF", needed: int = 40) -> None:
    """남은 페이지 여백이 needed mm 미만이면 새 페이지"""
    if pdf.get_y() > pdf.eph - needed:
        pdf.add_page()


def run_reporter(estimate_data: dict) -> str | None:
    """PDF 견적서 생성 후 base64 반환, 실패 시 None"""
    if not _FPDF_OK:
        logger.error("fpdf2 미설치 — PDF 생성 불가. pip install fpdf2 필요.")
        return None

    font_path = _find_font()
    if not font_path:
        logger.error(
            "한글 폰트 없음 — PDF 생성 중단. Railway 환경에서는 Dockerfile에 "
            "RUN apt-get install -y fonts-nanum 추가 필요. 후보 경로: %s",
            _FONT_REGULAR_CANDIDATES,
        )
        return None

    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_margins(15, 15, 15)

        bold_path = _find_bold_font()
        pdf.add_font("KR", "", font_path)
        if bold_path:
            pdf.add_font("KR", "B", bold_path)
        else:
            pdf.add_font("KR", "B", font_path)
            bold_path = font_path
        font_name = "KR"
        logger.info("한글 폰트 로드 성공: regular=%s bold=%s", font_path, bold_path)
        _bold_style = "B"

        # ── 헤더 ──────────────────────────────────────
        pdf.set_font(font_name, _bold_style, 20)
        pdf.cell(0, 14, "JH EstimateAI  견적서", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.set_font(font_name, "", 9)
        pdf.set_text_color(140, 140, 140)
        pdf.cell(0, 6, f"작성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  JH EstimateAI powered by Claude",
                 new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(6)

        # ── 고객 정보 (v2 optional) ──────────────────
        customer_name = estimate_data.get("customer_name")
        address = estimate_data.get("address")
        inquiry_id = estimate_data.get("inquiry_id")
        if customer_name or address or inquiry_id:
            pdf.set_font(font_name, _bold_style, 12)
            pdf.cell(0, 8, "■ 견적 요청 정보", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(font_name, "", 10)
            meta_rows = []
            if inquiry_id:     meta_rows.append(("요청번호", inquiry_id))
            if customer_name:  meta_rows.append(("고객명", customer_name))
            if address:        meta_rows.append(("현장 주소", address))
            for label, value in meta_rows:
                pdf.set_font(font_name, _bold_style, 10)
                pdf.cell(50, 7, label, border="B")
                pdf.set_font(font_name, "", 10)
                pdf.cell(0, 7, str(value), border="B", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(4)

        # ── 공사 개요 ─────────────────────────────────
        pdf.set_font(font_name, _bold_style, 12)
        pdf.cell(0, 8, "■ 공사 개요", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(font_name, "", 10)

        type_ = estimate_data.get("type", "")
        area = estimate_data.get("area", 0)
        min_cost = estimate_data.get("min_cost", 0)
        max_cost = estimate_data.get("max_cost", 0)
        unit_price = estimate_data.get("unit_price", 0)

        info_rows = [
            ("공사 유형", str(type_)),
            ("전체 면적", f"{area} m²"),
            ("m² 당 기준 단가", f"{unit_price:,} 원" if unit_price else "-"),
            ("최소 견적", f"{min_cost:,} 원" if min_cost else "-"),
            ("최대 견적", f"{max_cost:,} 원" if max_cost else "-"),
        ]
        for label, value in info_rows:
            pdf.set_font(font_name, _bold_style, 10)
            pdf.cell(50, 7, label, border="B")
            pdf.set_font(font_name, "", 10)
            pdf.cell(0, 7, value, border="B", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(6)

        # ── 공종별 내역 테이블 ─────────────────────────
        _check_page(pdf, 60)
        pdf.set_font(font_name, _bold_style, 12)
        pdf.cell(0, 8, "■ 공종별 견적 내역", new_x="LMARGIN", new_y="NEXT")

        breakdown = estimate_data.get("breakdown", {})
        total = sum(breakdown.values()) or 1

        pdf.set_fill_color(45, 45, 60)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font(font_name, _bold_style, 10)
        pdf.cell(80, 8, "공종", border=1, fill=True)
        pdf.cell(55, 8, "금액", border=1, fill=True, align="R")
        pdf.cell(35, 8, "비중", border=1, fill=True, align="R", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)

        fill = False
        pdf.set_font(font_name, "", 10)
        for category, amount in breakdown.items():
            ratio = f"{amount / total * 100:.1f}%" if total > 0 else "0%"
            pdf.set_fill_color(245, 245, 250) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.cell(80, 7, category, border=1, fill=True)
            pdf.cell(55, 7, f"{amount:,} 원", border=1, fill=True, align="R")
            pdf.cell(35, 7, ratio, border=1, fill=True, align="R", new_x="LMARGIN", new_y="NEXT")
            fill = not fill

        pdf.set_fill_color(220, 230, 255)
        pdf.set_font(font_name, _bold_style, 11)
        pdf.cell(80, 9, "합 계 (기준가)", border=1, fill=True)
        pdf.cell(55, 9, f"{total:,} 원", border=1, fill=True, align="R")
        pdf.cell(35, 9, "100%", border=1, fill=True, align="R", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(6)

        # ── 3단계 비교 견적 (enriched) ────────────────
        tiers = estimate_data.get("tiers")
        if tiers:
            _check_page(pdf, 50)
            pdf.set_font(font_name, _bold_style, 12)
            pdf.cell(0, 8, "■ 3단계 비교 견적", new_x="LMARGIN", new_y="NEXT")

            pdf.set_fill_color(45, 45, 60)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font(font_name, _bold_style, 10)
            pdf.cell(38, 8, "등급", border=1, fill=True)
            pdf.cell(52, 8, "최소 견적", border=1, fill=True, align="R")
            pdf.cell(52, 8, "최대 견적", border=1, fill=True, align="R")
            pdf.cell(28, 8, "권장", border=1, fill=True, align="C", new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)

            for key, label in [("budget", "저가형"), ("standard", "표준형"), ("premium", "고급형")]:
                t = tiers.get(key, {})
                is_rec = bool(t.get("recommended"))
                t_min = t.get("min", 0) or 0
                t_max = t.get("max", 0) or 0
                if is_rec:
                    pdf.set_fill_color(220, 235, 255)
                else:
                    pdf.set_fill_color(255, 255, 255)
                pdf.set_font(font_name, _bold_style if is_rec else "", 10)
                pdf.cell(38, 8, label, border=1, fill=True)
                pdf.set_font(font_name, "", 10)
                pdf.cell(52, 8, f"{t_min:,} 원", border=1, fill=True, align="R")
                pdf.cell(52, 8, f"{t_max:,} 원", border=1, fill=True, align="R")
                pdf.cell(28, 8, "★ 권장" if is_rec else "", border=1, fill=True, align="C",
                         new_x="LMARGIN", new_y="NEXT")
            pdf.ln(6)

        # ── 포함·별도·제외 항목 (enriched) ───────────
        inclusions = estimate_data.get("inclusions")
        if inclusions:
            _check_page(pdf, 40)
            pdf.set_font(font_name, _bold_style, 12)
            pdf.cell(0, 8, "■ 포함·별도·제외 항목", new_x="LMARGIN", new_y="NEXT")

            section_cfg = [
                ("included", "포함 항목", (34, 197, 94)),
                ("separate", "별도 항목", (200, 150, 0)),
                ("excluded", "제외 항목", (140, 140, 155)),
            ]
            for key, label, color in section_cfg:
                items = inclusions.get(key, []) or []
                if not items:
                    continue
                _check_page(pdf, 20)
                pdf.set_font(font_name, _bold_style, 10)
                pdf.set_text_color(*color)
                pdf.cell(0, 7, f"  {label} ({len(items)}건)", new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(0, 0, 0)
                pdf.set_font(font_name, "", 9)
                for item in items:
                    pdf.cell(8, 6, "·")
                    pdf.cell(0, 6, str(item), new_x="LMARGIN", new_y="NEXT")
                pdf.ln(2)
            pdf.ln(4)

        # ── 공사 일정 (enriched) ──────────────────────
        schedule = estimate_data.get("schedule")
        if schedule:
            _check_page(pdf, 45)
            pdf.set_font(font_name, _bold_style, 12)
            pdf.cell(0, 8, "■ 공사 일정", new_x="LMARGIN", new_y="NEXT")

            rec_days = schedule.get("recommendedDays", 0)
            risk = schedule.get("risk") or {}
            start = schedule.get("preferredStartDate", "") or ""
            end = schedule.get("preferredEndDate", "") or ""

            sched_rows = [("권장 공사일수", f"{rec_days}일")]
            if start:
                sched_rows.append(("희망 착공일", start))
            if end:
                sched_rows.append(("희망 완료일", end))
            if risk.get("label"):
                sched_rows.append(("일정 평가", risk["label"]))

            for label, value in sched_rows:
                pdf.set_font(font_name, _bold_style, 10)
                pdf.cell(50, 7, label, border="B")
                pdf.set_font(font_name, "", 10)
                pdf.cell(0, 7, value, border="B", new_x="LMARGIN", new_y="NEXT")

            if risk.get("message"):
                pdf.set_font(font_name, "", 9)
                pdf.set_text_color(120, 120, 140)
                pdf.ln(2)
                pdf.multi_cell(0, 5, risk["message"], new_x="LMARGIN", new_y="NEXT")
                pdf.set_text_color(0, 0, 0)
            pdf.ln(6)

        # ── 공정 계획 (enriched) ──────────────────────
        phases = (schedule or {}).get("phases", []) if schedule else []
        if phases:
            _check_page(pdf, 50)
            pdf.set_font(font_name, _bold_style, 12)
            pdf.cell(0, 8, "■ 공정 계획", new_x="LMARGIN", new_y="NEXT")

            pdf.set_fill_color(45, 45, 60)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font(font_name, _bold_style, 9)
            pdf.cell(55, 7, "공정", border=1, fill=True)
            pdf.cell(18, 7, "일수", border=1, fill=True, align="C")
            pdf.cell(97, 7, "주요 작업", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)

            fill = False
            pdf.set_font(font_name, "", 9)
            for phase in phases:
                tasks = phase.get("tasks", []) or []
                tasks_str = ", ".join(str(t) for t in tasks[:3])
                days_val = phase.get("days", 0) or 0
                if fill:
                    pdf.set_fill_color(245, 245, 250)
                else:
                    pdf.set_fill_color(255, 255, 255)
                _check_page(pdf, 15)
                pdf.cell(55, 7, str(phase.get("name", "")), border=1, fill=True)
                pdf.cell(18, 7, f"{days_val}일", border=1, fill=True, align="C")
                pdf.cell(97, 7, tasks_str, border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
                fill = not fill
            pdf.ln(6)

        # ── 전문가 검증 결과 ──────────────────────────
        flags = estimate_data.get("validator_flags", [])
        expert_comment = estimate_data.get("expert_comment", "")

        if flags:
            _check_page(pdf, 40)
            pdf.set_font(font_name, _bold_style, 12)
            pdf.cell(0, 8, "■ 전문가 검증 결과", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(font_name, "", 9)
            for flag in flags:
                severity = flag.get("severity", "warning")
                color = (220, 50, 50) if severity == "error" else (200, 140, 0)
                label = "[오류]" if severity == "error" else "[주의]"
                suggestion = flag.get("suggestion", "")
                msg_line = f"{flag.get('category', '')} - {flag.get('message', '')}"
                if suggestion:
                    msg_line += f"\n    >> {suggestion}"
                pdf.set_font(font_name, _bold_style, 9)
                pdf.set_text_color(*color)
                pdf.cell(20, 6, label)
                pdf.set_text_color(0, 0, 0)
                pdf.set_font(font_name, "", 9)
                pdf.multi_cell(0, 6, msg_line, new_x="LMARGIN", new_y="NEXT")
                pdf.ln(1)
            pdf.ln(3)

        if expert_comment:
            _check_page(pdf, 30)
            pdf.set_font(font_name, _bold_style, 12)
            pdf.cell(0, 8, "■ 전문가 총평", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(font_name, "", 10)
            pdf.multi_cell(0, 6, expert_comment)
            pdf.ln(4)

        # ── AI 실무 코멘트 (enriched) ─────────────────
        commentary = estimate_data.get("commentary", []) or []
        if commentary:
            _check_page(pdf, 40)
            pdf.set_font(font_name, _bold_style, 12)
            pdf.cell(0, 8, "■ AI 실무 코멘트", new_x="LMARGIN", new_y="NEXT")
            for i, comment in enumerate(commentary, 1):
                _check_page(pdf, 15)
                pdf.set_font(font_name, _bold_style, 10)
                pdf.cell(8, 6, f"{i}.")
                pdf.set_font(font_name, "", 10)
                pdf.multi_cell(0, 6, str(comment), new_x="LMARGIN", new_y="NEXT")
                pdf.ln(2)
            pdf.ln(4)

        # ── AI 요약 ────────────────────────────────────
        summary = estimate_data.get("summary", "")
        if summary:
            _check_page(pdf, 30)
            pdf.set_font(font_name, _bold_style, 12)
            pdf.cell(0, 8, "■ AI 견적 요약", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(font_name, "", 10)
            pdf.multi_cell(0, 6, summary)
            pdf.ln(4)

        # ── 주의사항 ──────────────────────────────────
        _check_page(pdf, 20)
        pdf.set_font(font_name, "", 8)
        pdf.set_text_color(160, 160, 160)
        pdf.multi_cell(0, 5,
            "※ 본 견적서는 AI 자동 산출 결과입니다. 현장 실측 및 자재 선정 후 최종 금액이 달라질 수 있습니다.\n"
            "※ JH EstimateAI는 18년 현장 경험 기반 데이터를 활용하나, 최종 계약 전 전문가 검토를 권장합니다."
        )

        pdf_bytes = pdf.output()
        return base64.b64encode(pdf_bytes).decode("utf-8")

    except Exception as e:
        logger.error("run_reporter 실패: %s", e, exc_info=True)
        return None


def run_reporter_excel(estimate_data: dict) -> str | None:
    """Excel 견적서 생성 후 base64 반환, 실패 시 None"""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        return None

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "견적서"

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22

        purple = "7C6AF7"
        green = "22D3A0"
        gray = "A09EB8"
        red = "F87171"
        amber = "FBBF24"
        border_thin = Border(
            left=Side(style='thin', color='444444'),
            right=Side(style='thin', color='444444'),
            top=Side(style='thin', color='444444'),
            bottom=Side(style='thin', color='444444'),
        )

        def section_header(ws, row, text, color=purple):
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = text
            ws[f'A{row}'].font = Font(name='맑은 고딕', bold=True, size=11, color=color)
            ws.row_dimensions[row].height = 20
            return row + 1

        def kv_row(ws, row, label, value, label_color=gray, value_color='E8E6F0'):
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name='맑은 고딕', bold=True, color=label_color)
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'] = str(value)
            ws[f'B{row}'].font = Font(name='맑은 고딕', color=value_color)
            return row + 1

        row = 1

        # 제목
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'JH EstimateAI 견적서'
        ws[f'A{row}'].font = Font(name='맑은 고딕', size=18, bold=True, color=purple)
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 36
        row += 1

        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = '18년 현장 경험 기반 · 5 에이전트 AI 견적 시스템'
        ws[f'A{row}'].font = Font(name='맑은 고딕', size=10, color=gray)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # 고객 정보
        meta_rows = []
        if estimate_data.get('inquiry_id'):    meta_rows.append(('요청번호', estimate_data['inquiry_id']))
        if estimate_data.get('customer_name'): meta_rows.append(('고객명', estimate_data['customer_name']))
        if estimate_data.get('address'):       meta_rows.append(('현장 주소', estimate_data['address']))
        if meta_rows:
            row = section_header(ws, row, '■ 견적 요청 정보')
            for label, value in meta_rows:
                row = kv_row(ws, row, label, value)
            row += 1

        # 기본 정보
        row = section_header(ws, row, '■ 공사 개요')
        info = [
            ('공사 유형', estimate_data.get('type', '-')),
            ('면적', f"{estimate_data.get('area', 0)}m²"),
            ('최소 견적', f"{estimate_data.get('min_cost', 0):,}원"),
            ('최대 견적', f"{estimate_data.get('max_cost', 0):,}원"),
            ('기준 단가', f"{estimate_data.get('unit_price', 0):,}원/m²"),
            ('작성일', datetime.now().strftime('%Y-%m-%d')),
        ]
        for label, value in info:
            row = kv_row(ws, row, label, value)
        row += 1

        # 공종별 내역 헤더
        row = section_header(ws, row, '■ 공종별 견적 내역')
        headers = ['공종', '금액 (원)', '비율 (%)', '비고']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(name='맑은 고딕', bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=purple, end_color=purple, fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = border_thin
        row += 1

        breakdown = estimate_data.get('breakdown', {}) or {}
        total = sum(breakdown.values()) or 1
        for name, amount in breakdown.items():
            pct = round(amount / total * 100, 1)
            ws.cell(row=row, column=1, value=name).border = border_thin
            amt_cell = ws.cell(row=row, column=2, value=amount)
            amt_cell.number_format = '#,##0'
            amt_cell.border = border_thin
            pct_cell = ws.cell(row=row, column=3, value=pct)
            pct_cell.number_format = '0.0"%"'
            pct_cell.border = border_thin
            ws.cell(row=row, column=4, value='').border = border_thin
            row += 1

        ws.cell(row=row, column=1, value='합계').font = Font(name='맑은 고딕', bold=True)
        total_cell = ws.cell(row=row, column=2, value=total)
        total_cell.font = Font(name='맑은 고딕', bold=True, color=purple)
        total_cell.number_format = '#,##0'
        row += 2

        # 3단계 비교견적 (enriched)
        tiers = estimate_data.get('tiers')
        if tiers:
            row = section_header(ws, row, '■ 3단계 비교견적')
            tier_headers = ['등급', '최소 견적 (원)', '최대 견적 (원)', '권장']
            for col, h in enumerate(tier_headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(name='맑은 고딕', bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2D2D3C', end_color='2D2D3C', fill_type='solid')
                cell.border = border_thin
            row += 1
            for key, label in [('budget', '저가형'), ('standard', '표준형'), ('premium', '고급형')]:
                t = tiers.get(key, {}) or {}
                is_rec = bool(t.get('recommended'))
                t_min = t.get('min', 0) or 0
                t_max = t.get('max', 0) or 0
                fill_color = 'DCF0FF' if is_rec else 'FFFFFF'
                for col, val in enumerate([label, t_min, t_max, '★ 권장' if is_rec else ''], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = Font(name='맑은 고딕', bold=is_rec)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    cell.border = border_thin
                    if col in (2, 3):
                        cell.number_format = '#,##0'
                row += 1
            row += 1

        # 포함·별도·제외 항목 (enriched)
        inclusions = estimate_data.get('inclusions')
        if inclusions:
            row = section_header(ws, row, '■ 포함·별도·제외 항목')
            for key, label, color in [
                ('included', '포함 항목', '22C55E'),
                ('separate', '별도 항목', 'CA8A04'),
                ('excluded', '제외 항목', '9CA3AF'),
            ]:
                items = inclusions.get(key, []) or []
                if not items:
                    continue
                ws.merge_cells(f'A{row}:D{row}')
                ws[f'A{row}'] = f'  {label} ({len(items)}건)'
                ws[f'A{row}'].font = Font(name='맑은 고딕', bold=True, color=color)
                row += 1
                for item in items:
                    ws.merge_cells(f'A{row}:D{row}')
                    ws[f'A{row}'] = f'  · {item}'
                    ws[f'A{row}'].font = Font(name='맑은 고딕', size=9, color='C4C2D8')
                    row += 1
            row += 1

        # 공사 일정 (enriched)
        schedule = estimate_data.get('schedule')
        if schedule:
            row = section_header(ws, row, '■ 공사 일정')
            rec_days = schedule.get('recommendedDays', 0) or 0
            risk = schedule.get('risk') or {}
            sched_items = [('권장 공사일수', f'{rec_days}일')]
            if schedule.get('preferredStartDate'):
                sched_items.append(('희망 착공일', schedule['preferredStartDate']))
            if schedule.get('preferredEndDate'):
                sched_items.append(('희망 완료일', schedule['preferredEndDate']))
            if risk.get('label'):
                sched_items.append(('일정 평가', risk['label']))
            if risk.get('message'):
                sched_items.append(('일정 메시지', risk['message']))
            for label, value in sched_items:
                row = kv_row(ws, row, label, value)
            row += 1

        # 공정 계획 (enriched)
        phases = (schedule or {}).get('phases', []) if schedule else []
        if phases:
            row = section_header(ws, row, '■ 공정 계획')
            phase_headers = ['공정명', '일수', '주요 작업', '비고']
            for col, h in enumerate(phase_headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(name='맑은 고딕', bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2D2D3C', end_color='2D2D3C', fill_type='solid')
                cell.border = border_thin
            row += 1
            for phase in phases:
                tasks = phase.get('tasks', []) or []
                tasks_str = ', '.join(str(t) for t in tasks[:3])
                for col, val in enumerate([
                    phase.get('name', ''),
                    f"{phase.get('days', 0) or 0}일",
                    tasks_str,
                    '',
                ], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = Font(name='맑은 고딕', size=9)
                    cell.border = border_thin
                    cell.alignment = Alignment(wrap_text=True)
                row += 1
            row += 1

        # AI 실무 코멘트 (enriched)
        commentary = estimate_data.get('commentary', []) or []
        if commentary:
            row = section_header(ws, row, '■ AI 실무 코멘트', color=green)
            for i, comment in enumerate(commentary, 1):
                ws.merge_cells(f'A{row}:D{row}')
                ws[f'A{row}'] = f'{i}. {comment}'
                ws[f'A{row}'].font = Font(name='맑은 고딕', size=9, color='C4C2D8')
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 30
                row += 1
            row += 1

        # AI 요약
        if estimate_data.get('summary'):
            row = section_header(ws, row, '■ AI 전문가 요약', color=green)
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = estimate_data['summary']
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 45
            row += 2

        # VALIDATOR 플래그
        flags = estimate_data.get('validator_flags', [])
        if flags:
            row = section_header(ws, row, '■ VALIDATOR — 18년 현장 검증 결과', color=red)
            for flag in flags:
                sev = flag.get('severity', 'warning')
                msg = f"[{flag.get('category','')}] {flag.get('message','')} → {flag.get('suggestion','')}"
                ws.merge_cells(f'A{row}:D{row}')
                ws[f'A{row}'] = f"{'⚠ ' if sev=='warning' else '✕ '}{msg}"
                ws[f'A{row}'].font = Font(name='맑은 고딕', color=amber if sev == 'warning' else red)
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 30
                row += 1

        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode()

    except Exception as e:
        logger.error("run_reporter_excel 실패: %s", e, exc_info=True)
        return None
